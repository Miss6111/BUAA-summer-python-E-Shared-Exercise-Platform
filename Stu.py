import openpyxl
import sympy

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import sqlalchemy.orm.session  # 数据库操作核心
from sqlalchemy import desc
from sqlalchemy.ext.declarative import declarative_base  # 父类
from sqlalchemy import or_
from datetime import datetime
from sqlalchemy import DateTime

Base = sqlalchemy.orm.declarative_base()
DB_connect = 'mysql+mysqldb://root:86901260@localhost/mydb_1'
engine = create_engine(DB_connect, echo=False)


# 评论表
class Comments(Base):
    __tablename__ = 'comments'
    id = sqlalchemy.Column(sqlalchemy.BIGINT, primary_key=True)
    sender = sqlalchemy.Column(sqlalchemy.Integer)  # 评论的userid
    qid = sqlalchemy.Column(sqlalchemy.Integer)  # 对哪个问题评论
    content = sqlalchemy.Column(sqlalchemy.String(200))  # 评论内容


# 收藏问题_stu
class Star_stu(Base):
    __tablename__ = 'star_stu'
    uid = sqlalchemy.Column(sqlalchemy.BIGINT, sqlalchemy.ForeignKey("stus.uid"), primary_key=True)
    qid = sqlalchemy.Column(sqlalchemy.BIGINT, sqlalchemy.ForeignKey("questions.qid"), primary_key=True)


# 多对多表
class Chap_ques(Base):  # 存问题-章节
    __tablename__ = 'chap_ques'
    name = sqlalchemy.Column(sqlalchemy.String(20), sqlalchemy.ForeignKey("chapters.name"), primary_key=True)
    qid = sqlalchemy.Column(sqlalchemy.BIGINT, sqlalchemy.ForeignKey("questions.qid"), primary_key=True)


class Ques_qgroup(Base):  # 存每个问题---问题组
    __tablename__ = 'ques_qgroup'
    gid = sqlalchemy.Column(sqlalchemy.BIGINT, sqlalchemy.ForeignKey("qgroups.gid"), primary_key=True)
    qid = sqlalchemy.Column(sqlalchemy.BIGINT, sqlalchemy.ForeignKey("questions.qid"), primary_key=True)


class Stu_group(Base):  # 学生--学生小组
    __tablename__ = 'stu_group'
    uid = sqlalchemy.Column(sqlalchemy.BIGINT, sqlalchemy.ForeignKey("stus.uid"), primary_key=True)
    gid = sqlalchemy.Column(sqlalchemy.BIGINT, sqlalchemy.ForeignKey("groups.gid"), primary_key=True)


class Stu_qgroup(Base):  # 学生--问题组
    __tablename__ = 'stu_qgroup'
    uid = sqlalchemy.Column(sqlalchemy.BIGINT, sqlalchemy.ForeignKey("stus.uid"), primary_key=True)
    gid = sqlalchemy.Column(sqlalchemy.BIGINT, sqlalchemy.ForeignKey("qgroups.gid"), primary_key=True)


class Ggroup_group(Base):  # 存问题组--学生组
    __tablename__ = 'qgroup_group'
    qgid = sqlalchemy.Column(sqlalchemy.BIGINT, sqlalchemy.ForeignKey("qgroups.gid"), primary_key=True)
    gid = sqlalchemy.Column(sqlalchemy.BIGINT, sqlalchemy.ForeignKey("groups.gid"), primary_key=True)


class Chapters(Base):  # 有哪些章节
    __tablename__ = 'chapters'
    name = sqlalchemy.Column(sqlalchemy.String(20), primary_key=True)
    ques = sqlalchemy.orm.relationship("Questions", secondary="chap_ques", backref="Chapters",
                                       cascade='all')  # 这个章节对应的所有题


class Questions(Base):  # 有哪些问题
    __tablename__ = 'questions'
    qid = sqlalchemy.Column(sqlalchemy.BIGINT, primary_key=True)  # 当前问题的编号
    uid = sqlalchemy.Column(sqlalchemy.Integer)  # 创建人的id
    name = sqlalchemy.Column(sqlalchemy.String(20))  # 题目
    title = sqlalchemy.Column(sqlalchemy.String(1000))  # 题干
    chapter = sqlalchemy.Column(sqlalchemy.String(10))  # 章节
    # 所属问题组
    qgroups = sqlalchemy.orm.relationship("QGroups", secondary="ques_qgroup", backref="Questions", cascade='all')
    # 题型,0选择，1填空
    type = sqlalchemy.Column(sqlalchemy.Integer)
    # ”1010“为选AC
    answer = sqlalchemy.Column(sqlalchemy.String(5))
    # ABCD四个选项，填空默认显示A
    answerA = sqlalchemy.Column(sqlalchemy.String(100))
    answerB = sqlalchemy.Column(sqlalchemy.String(100))
    answerC = sqlalchemy.Column(sqlalchemy.String(100))
    answerD = sqlalchemy.Column(sqlalchemy.String(100))
    # 填空题答案
    gap = sqlalchemy.Column(sqlalchemy.String(500))
    public = sqlalchemy.Column(sqlalchemy.Boolean)  # 是否是所有人可见
    total = sqlalchemy.Column(sqlalchemy.Integer)
    right = sqlalchemy.Column(sqlalchemy.Integer)


class Stus(Base):
    __tablename__ = 'stus'  # 名字
    uid = sqlalchemy.Column(sqlalchemy.BIGINT, primary_key=True)
    # 进行属性与字段映射，名字可以不一
    name = sqlalchemy.Column(sqlalchemy.String(50))

    password = sqlalchemy.Column(sqlalchemy.String(50))
    # 学习小组，学习小组由管理员建立，用户可以通过学习小组的名字加入学习小组
    groups = sqlalchemy.orm.relationship("Groups", secondary="stu_group", backref="Stus", cascade='all')
    # 共享问题组信息，当用户加入一个学习小组，该用户就可以看到学习小组中成员在小组内共享的所有问题组，qgroups就保存了用户可见的所有问题组
    qgroups = sqlalchemy.orm.relationship("QGroups", secondary="stu_qgroup", backref="Stus", cascade='all')
    # 收藏的问题
    starquestions = sqlalchemy.orm.relationship("Questions", secondary="star_stu", backref="Stus", cascade='all')
    # 格言
    quote = sqlalchemy.Column(sqlalchemy.String(100))
    # 简介
    Bi = sqlalchemy.Column(sqlalchemy.String(100))
    # 是否是管理员
    issuper = sqlalchemy.Column(sqlalchemy.Boolean)


class Groups(Base):  # 用户小组
    __tablename__ = 'groups'
    gid = sqlalchemy.Column(sqlalchemy.BIGINT, primary_key=True)  # 组id
    name = sqlalchemy.Column(sqlalchemy.String(50))  # 组名
    uid = sqlalchemy.Column(sqlalchemy.Integer)  # 创造者的id
    qgroups = sqlalchemy.orm.relationship("QGroups", secondary="qgroup_group", backref="Groups")  # 这个用户小组拥有权限的问题小组


class QGroups(Base):  # 问题小组，由每个用户主动创建
    __tablename__ = 'qgroups'
    gid = sqlalchemy.Column(sqlalchemy.BIGINT, primary_key=True, autoincrement=True)
    uid = sqlalchemy.Column(sqlalchemy.Integer)  # 创造者
    name = sqlalchemy.Column(sqlalchemy.String(20), primary_key=True)
    public = sqlalchemy.Column(sqlalchemy.Boolean)  # 是否公开


# 做题记录
class Records(Base):
    __tablename__ = 'records'
    uid = sqlalchemy.Column(sqlalchemy.BIGINT, primary_key=True)
    qid = sqlalchemy.Column(sqlalchemy.BIGINT, primary_key=True)
    right = sqlalchemy.Column(sqlalchemy.Boolean, primary_key=True)
    time = sqlalchemy.Column(DateTime, default=datetime.now)
    rank = sqlalchemy.Column(sqlalchemy.BIGINT, primary_key=True)
    rate = sqlalchemy.Column(sqlalchemy.BIGINT, primary_key=True)
    never = sqlalchemy.Column(sqlalchemy.Boolean, primary_key=True)  # 当本题已经连续作对三次则不再推荐


def create_session():  # session用来操作数据库
    """

    :return:
    """
    session_ = sessionmaker(bind=engine)  # 一个session是一个对数据库链接的包装
    session = session_()  # 实例化session对象
    return session


def check_name(C, name):
    """

    :param C:
    :param name:
    :return:
    """
    session = create_session()
    if len(session.query(C).filter(C.name == name).all()) == 0:
        session.close()
        return True
    session.close()
    return False


# 任务一，个人信息管理
def create_new_user(name, password, manager):  # 按下注册确定按键的瞬间,创建新用户
    """

    :param name:
    :param password:
    :param manager:
    :return:
    """
    s = create_session()
    if not check_name(Stus, name):
        s.close()
        return False
    else:
        new = Stus(name=name, password=password, issuper=manager, Bi="你还没有写任何简介", quote="", groups=[],
                   qgroups=[])
        s.add(new)
        s.commit()
        s.close()
        return True


def change_password(password, name):  # 改密码
    """

    :param password:
    """
    s = create_session()
    s.query(Stus).filter(Stus.name == name).first().password = password
    s.commit()
    s.close()


def change_name(new, name):  # 改名字，按下确定瞬间
    """

    :param new:
    :return:
    """
    if not check_name(Stus, new):
        return False
    else:
        s = create_session()
        stu = s.query(Stus).filter(Stus.name == name).first()
        stu.name = new
        s.commit()
        s.close()
        return True


# 当前用户属于的所有组
def show_users_groups(name):
    s = create_session()
    groups = []
    for i in s.query(Stus).filter(Stus.name == name).first().groups:
        groups.append(i.name)
    s.close()
    return groups


def change_quote(new, name):  # 改格言
    """

    :param new:
    """
    s = create_session()
    s.query(Stus).filter(Stus.name == name).first().quote = new
    s.commit()
    s.close()


def change_bi(new, name):  # 改简介
    """

    :param new:
    """
    s = create_session()
    s.query(Stus).filter(Stus.name == name).first().Bi = new
    s.commit()
    s.close()


def login(name, password):  # 登入瞬间
    """

    :param name:
    :param password:
    :return:
    """
    s = create_session()
    if check_name(Stus, name):  # 没这个账号名
        return False
    else:
        passw = s.query(Stus).filter(Stus.name == name).first().password
        if passw == password:  # 账号密码都正确
            s.close()
            return True
        else:
            s.close()
            return False


# 任务二 管理员创建小组，将他人加入小组，用户搜索加入小组，注意，此时要更新学生的问题组权限
def check_super(name):
    """

    :return:
    """
    s = create_session()
    ans = s.query(Stus).filter(Stus.name == name).first().issuper
    s.close()
    return ans


def all_groups():  # 管理员界面使用，目前的所有用户小组
    groups = []
    s = create_session()
    all = s.query(Groups).all()
    for i in all:
        groups.append(i.name)
    return groups


#
#
#
#
def create_new_group(g_name, name):  # 创建一个空的新的小组(小组名)
    """

    :param g_name:
    :return:
    """
    s = create_session()
    if not check_name(Groups, g_name):  # 名字存在
        s.close()
        return False
    else:
        uid = s.query(Stus).filter(Stus.name == name).first().uid
        new = Groups(name=g_name, uid=uid)
        s.add(new)
        s.commit()
        s.close()
        return True


# 将一组人加进group，注意问题权限
def add_into_group(users, g_name):  # 此处users为名字字符串数组
    """

    :param users:
    :param g_name:
    """
    s = create_session()
    g_name = g_name.strip()
    group = s.query(Groups).filter(Groups.name == g_name).first()  # 找到当前学生gruop

    for i in users:  # 遍历每一个学生
        stu = s.query(Stus).filter(Stus.name == i.strip()).first()
        stu.groups.append(group)  # 保证此时学生一定不在这个组

        for j in group.qgroups:  # 当前group的qgroups
            if not j in stu.qgroups:  # 这个学生目前的qgroups中不存在的才加入
                stu.qgroups.append(j)  # 学生加入权限
    s.commit()
    s.close()


def search_for_groups(gname, name):  # 搜索组用，除去了当前用户已经在的组
    s = create_session()
    gname = gname.strip()
    groups = s.query(Groups).filter(Groups.name.like('%' + gname + '%')).all()
    names = []
    uid = s.query(Stus).filter(Stus.name == name).first().uid
    temps = []
    for i in groups:
        inGroups = s.query(Stu_group).filter(Stu_group.uid == uid).all()
        for j in inGroups:
            temps.append(j.gid)
        if not i.gid in temps:
            names.append(i.name)
    return names
    s.close()


def search_students(gname, name):  # 去除了已经在表里的人
    """

    :param gname:
    :param name:
    :return:
    """
    s = create_session()
    gname = gname.strip()
    gid = s.query(Groups).filter(Groups.name == gname).first().gid
    uids_ = s.query(Stu_group).filter(Stu_group.gid == gid).all()
    uids = []
    for i in uids_:
        uids.append(i.uid)
    students = s.query(Stus).filter(Stus.name.like('%' + name + '%')).all()
    unames = []
    for i in students:
        if not i.uid in uids:
            unames.append(i.name)
    # 符合要求所有学生名字
    s.commit()
    s.close()
    return unames


# def delete_from_group(users, gname):  # 将部分人从组里删除
#     """
#
#     :param users:
#     :param gname:
#     """
#     s = create_session()
#     group = s.query(Groups).filter(Groups.name == gname).first()
#     qgroups = group.qgroups
#     for i in users:
#         user = s.query(Stus).filter(Stus.name == i).first()
#         user.groups.remove(group)
#         for j in qgroups:
#             user.groups.remove(j)
#     s.commit()
#     s.close()


def search_groups(page):  # 用户查找组时
    """

    :param page:
    :return:
    """
    # 分页显示，每页十条
    s = create_session()
    groups = s.query(Groups).limit(10).offset((page - 1) * 10).all()
    gnames = []
    for i in groups:
        gnames.append(i.name)
    s.close()
    # 返回值是组名数组
    return gnames


def user_add_into_group(gnames, name):  # 用户主动申请加入一串组,此时保证用户都不在这些组里
    """

    :param name:
    :param gnames:
    :return:
    """
    s = create_session()
    # 如果已经在组里，加入失败
    groups = s.query(Groups).filter(Groups.name.in_(gnames)).all()
    stu = s.query(Stus).filter(Stus.name == name).first()
    for i in groups:
        stu.groups.append(i)  # 关联的是整个而不是一个值
        print(i.name)
    for group in groups:
        qgroups = group.qgroups  # 当前group的qgroups
        # 这个学生目前的qgroups中不存在的才加入
        for j in qgroups:
            print(j.name)
            if not j in stu.qgroups:
                stu.qgroups.append(j)  # 学生加入权限
    s.commit()
    s.close()


# 任务三 上传 单个问题 或 一个文件的问题
def show_all_chapter():  # 返回所有的章节名字
    """

    :return:
    """
    s = create_session()
    name = []
    for i in s.query(Chapters).all():
        name.append(i.name)
    s.close()
    return name


def load_one_question(title, answer, chapter, my_type, answer1, answer2, answer3, answer4, gap, public, creater):
    """


    :param creater:
    :param title:
    :param answer:
    :param chapter:
    :param my_type:
    :param answer1:
    :param answer2:
    :param answer3:
    :param answer4:
    :param public:
    """
    s = create_session()
    c = s.query(Chapters).filter(Chapters.name == chapter).first()
    qid = len(s.query(Questions).all())
    q = Questions(qid=qid + 1, title=title, answer=answer, type=my_type, answerA=answer1, answerB=answer2,
                  answerC=answer3,
                  answerD=answer4, gap=gap, public=public, uid=s.query(Stus).filter(Stus.name == creater).first().uid,
                  total=0, right=0, chapter=chapter, name=title[:10])
    s.add(q)
    s.commit()
    c.ques.append(q)
    s.commit()
    s.close()
    # 分组，给问题加标签


def initial_data():
    """

    :param name:
    :param path:
    """
    create_new_user("manager", "666666", 1)
    s = create_session()
    for i in range(1, 9):
        s.add(Chapters(name='Chapter_' + str(i), ques=[]))
    s.commit()
    s.close()
    f = openpyxl.load_workbook("D:\\Users\\23673\\Desktop\\summer_python\\try.xlsx")  # 改成本地的地址
    names = f.sheetnames  # 所有sheet
    for sheet_name in names:  # 每一页
        sheet = f[sheet_name]
        rows = sheet.max_row
        for i in range(1, rows):  # 每一行是一个问题
            title = sheet.cell(i + 1, 1).value
            A = sheet.cell(i + 1, 2).value
            B = sheet.cell(i + 1, 3).value
            C = sheet.cell(i + 1, 4).value
            D = sheet.cell(i + 1, 5).value
            answer_ = sheet.cell(i + 1, 6).value
            chapters = ['Chapter_1', 'Chapter_2', 'Chapter_3', 'Chapter_4', 'Chapter_5', 'Chapter_6', 'Chapter_7',
                        'Chapter_8', 'Chapter_9']
            answer = ['0', '0', '0', '0']
            gap = ''
            if "A" in answer_:
                gap = A
                answer[0] = '1'
            if "B" in answer_:
                gap = B
                answer[1] = '1'
            if "C" in answer_:
                gap = C
                answer[2] = '1'
            if "D" in answer_:
                gap = D
                answer[3] = '1'
            if i % 2 == 0 or len(answer_) > 2:  # 选择
                if len(answer_) > 2:
                    load_one_question(title, ''.join(answer), chapters[int(i / 150) + 1], 2, A, B, C, D, '',
                                      public=True, creater='manager')
                # title, answer, chapter, my_type, answer1, answer2, answer3, answer4, gap, public, creater
                else:
                    load_one_question(title, ''.join(answer), chapters[int(i / 150) + 1], 0, A, B, C, D, '',
                                      public=True,
                                      creater='manager')
            else:  # 填空
                load_one_question(title, '', chapters[int(i / 150) + 1], 1, A, B, C, D, gap, public=True,
                                  creater='manager')


def load_files(path, name):  # 需要规定文件格式？？再想
    """

    :param name:
    :param path:
    """
    f = openpyxl.load_workbook(path)
    names = f.get_sheet_names()  # 所有sheet
    for sheet_name in names:  # 每一页
        sheet = f.get_sheet_by_name(sheet_name)
        rows = sheet.max_row
        for i in range(rows):  # 每一行是一个问题
            title = sheet.cell(i + 1, 1).value
            answer = sheet.cell(i + 1, 2).value
            chapter = sheet.cell(i + 1, 3).value
            mytype = sheet.cell(i + 1, 4).value
            answer1 = sheet.cell(i + 1, 5).value
            answer2 = sheet.cell(i + 1, 6).value
            answer3 = sheet.cell(i + 1, 7).value
            answer4 = sheet.cell(i + 1, 8).value
            # 默认非公开
            load_one_question(title=title, answer=answer, chapter=chapter, my_type=mytype,
                              answer1=answer1, answer2=answer2, answer3=answer3, answer4=answer4, public=False,
                              creater=name, gap=answer)


# 根据关键词搜索问题
def scope_questions(ques_name, chapters_name, mytype, user_name, qgroups):  # 关键词，章节，题型
    """

    :param ques_name:
    :param chapters_name:
    :param mytype:
    :param user_name:
    :return:
    """
    print([ques_name, chapters_name, mytype, user_name])
    s = create_session()
    # 搜索范围包括questions中的public或上传者为本人的，和qgroup中的
    uid = s.query(Stus).filter(Stus.name == user_name).first().uid
    stu = s.query(Stus).filter(Stus.name == user_name).first()
    if len(qgroups) == 0:
        qgroupids = [i.gid for i in stu.qgroups]
    else:
        qgroupids = [i.gid for i in qgroups]
    q = s.query(Questions).filter(Questions.chapter.in_(chapters_name)).filter(
        or_(Questions.type == mytype, Questions.type == 2 - mytype)).filter(
        Questions.title.like('%' + ques_name + '%')).all()
    ques = []
    for i in q:
        flag = False
        for j in i.qgroups:
            if j.gid in qgroupids:
                flag = True
                break
        if i.uid == uid or i.public == True or flag == True:
            ques.append(i.qid)
    # 目前仅支持关键词为title子串
    s.commit()
    s.close()
    return ques  # 返回值为满足要求的qid


# ************************************************************************************************************** #
# 根据关键词搜索问题 ["Chapter_1","Chapter_2"]
def scope_questions_title(ques_name, chapters_name, mytype, user_name, qgroups):  # 关键词，章节，题型
    """

    :param ques_name:
    :param chapters_name:
    :param mytype:
    :param user_name:
    :return:
    """
    print([ques_name, chapters_name, mytype, user_name])
    s = create_session()
    # 搜索范围包括questions中的public或上传者为本人的，和qgroup中的
    uid = s.query(Stus).filter(Stus.name == user_name).first().uid
    stu = s.query(Stus).filter(Stus.name == user_name).first()
    if len(qgroups) == 0:
        qgroupids = [i.gid for i in stu.qgroups]
    else:
        qgroupids = [i.gid for i in qgroups]
    q = s.query(Questions).filter(Questions.chapter.in_(chapters_name)).filter(
        or_(Questions.type == mytype, Questions.type == 2 - mytype)).filter(
        Questions.title.like('%' + ques_name + '%')).all()
    ques = []
    for i in q:
        flag = False
        for j in i.qgroups:
            if j.gid in qgroupids:
                flag = True
                break
        if i.uid == uid or i.public == True or flag == True:
            ques.append(i.title)
            print([i.type, i.chapter])
    # 目前仅支持关键词为title子串
    s.commit()
    s.close()
    return ques  # 返回值为满足要求的qid


def scope_questions_answer(ques_name, chapters_name, mytype, user_name):  # 关键词，章节，题型
    """

    :param ques_name:
    :param chapters_name:
    :param mytype:
    :param user_name:
    :return:
    """
    s = create_session()
    # 搜索范围包括questions中的public或上传者为本人的，和qgroup中的
    uid = s.query(Stus).filter(Stus.name == user_name).first().uid
    stu = s.query(Stus).filter(Stus.name == user_name).first()
    qgroupids = [i.gid for i in stu.qgroups]
    q = s.query(Questions).filter(Questions.chapter.in_(chapters_name)).filter(
        or_(Questions.type == mytype, Questions.type == 2 - mytype)).filter(
        Questions.title.like('%' + ques_name + '%')).all()
    ques = []
    for i in q:
        flag = False
        for j in i.qgroups:
            if j.gid in qgroupids:
                flag = True
                break
        if i.uid == uid or i.public == True or flag == True:
            ques.append(i.answer)
    # 目前仅支持关键词为title子串
    s.commit()
    s.close()
    return ques  # 返回值为满足要求的qid


def scope_questions_type(ques_name, chapters_name, mytype, user_name):  # 关键词，章节，题型
    """

    :param ques_name:
    :param chapters_name:
    :param mytype:
    :param user_name:
    :return:
    """
    s = create_session()
    # 搜索范围包括questions中的public或上传者为本人的，和qgroup中的
    uid = s.query(Stus).filter(Stus.name == user_name).first().uid
    stu = s.query(Stus).filter(Stus.name == user_name).first()
    qgroupids = [i.gid for i in stu.qgroups]
    q = s.query(Questions).filter(Questions.chapter.in_(chapters_name)).filter(
        or_(Questions.type == mytype, Questions.type == 2 - mytype)).filter(
        Questions.title.like('%' + ques_name + '%')).all()
    ques = []
    for i in q:
        flag = False
        for j in i.qgroups:
            if j.gid in qgroupids:
                flag = True
                break
        if i.uid == uid or i.public == True or flag == True:
            ques.append(i.type)
    # 目前仅支持关键词为title子串
    s.commit()
    s.close()
    return ques  # 返回值为满足要求的qid


def scope_questions_qid(ques_name, chapters_name, mytype, user_name):  # 关键词，章节，题型
    """

    :param ques_name:
    :param chapters_name:
    :param mytype:
    :param user_name:
    :return:
    """
    s = create_session()
    # 搜索范围包括questions中的public或上传者为本人的，和qgroup中的
    uid = s.query(Stus).filter(Stus.name == user_name).first().uid
    stu = s.query(Stus).filter(Stus.name == user_name).first()
    qgroupids = [i.gid for i in stu.qgroups]
    q = s.query(Questions).filter(Questions.chapter.in_(chapters_name)).filter(
        or_(Questions.type == mytype, Questions.type == 2 - mytype)).filter(
        Questions.title.like('%' + ques_name + '%')).all()
    ques = []
    for i in q:
        flag = False
        for j in i.qgroups:
            if j.gid in qgroupids:
                flag = True
                break
        if i.uid == uid or i.public == True or flag == True:
            ques.append(i.qid)
    # 目前仅支持关键词为title子串
    s.commit()
    s.close()
    return ques  # 返回值为满足要求的qid


# ************************************************************************************************************** #


# 问题共享功能
def create_own_ques_group(name, user_name):  # 某个用户可以选择构造一个问题组并命名，类比学生和学生组
    """

    :param name:
    :return:
    """
    if check_name(QGroups, name) == False:
        return False
    s = create_session()
    uid = s.query(Stus).filter(Stus.name == user_name).first().uid
    new = QGroups(name=name, uid=uid)
    # new = QGroups(name=name, uid=uid， public = False)
    s.add(new)
    s.commit()
    s.close()
    return True


def add_ques_into_group(name, questions):  # 传入问题编号/后续可以考虑改成名字
    """

    :param name:
    :param questions:
    """
    s = create_session()
    group = s.query(QGroups).filter(QGroups.name == name).first()
    print(name)
    for i in questions:  # 将选中问题加入问题组中
        temp = s.query(Questions).filter(Questions.title == i).first()
        temp.qgroups.append(group)
    s.commit()
    s.close()


def share_qgroup_with_group(qgname, gname):  # 与特定的用户组分享特定的问题组
    """

    :param qgname:
    :param gname:
    """
    s = create_session()
    gid = s.query(Groups).filter(Groups.name == gname).first().gid  # 得到这个用户组的gid
    temp = s.query(Stu_group).filter(Stu_group.gid == gid).all()
    ids = []  # 得到用户组的所有用户的uid
    for i in temp:
        ids.append(i.uid)
    stus = s.query(Stus).filter(Stus.uid.in_(ids)).all()  # 得到用户组所有用户

    qgroup = s.query(QGroups).filter(QGroups.name == qgname).first()  # 得到这个问题组

    group = s.query(Groups).filter(Groups.name == gname).first()
    group.qgroups.append(qgroup)

    for i in stus:
        if not qgroup in i.qgroups:
            i.qgroups.append(qgroup)  # 依次建立联系

    s.commit()
    s.close()


# 只关心当前评论显示在在哪个答案的下方
def send_comments(qid, content, user_name):
    """

    :param user_name:
    :param qid:
    :param content:
    """
    s = create_session()
    stu = s.query(Stus).filter(Stus.name == user_name).first()
    new = Comments(qid=qid, content=content, sender=stu.uid)  # 保证不能被改变
    s.add(new)
    s.commit()
    s.close()


# 点击加号显示所有的评论-->下画框
def show_some_comments(qid):
    """

    :return: 返回三条评论
    """
    s = create_session()
    re = []
    comments = s.query(Comments).filter(Comments.qid == qid).limit(30).all()
    for i in comments:
        sender = s.query(Stus).filter(Stus.uid == i.sender).first().name
        re.append([sender, i.content])
    s.close()
    return re  # 返回的是[内容，发送人]


# def show_more_comments(qid):
#     """
#    # 全部评论
#     :return: 返回全部评论
#     """
#     s = create_session()
#     re = []
#     comments = s.query(Comments).all()
#     for i in comments:
#         sender = s.query(Stus).filter(Stus.uid == i.sender).filter(Comments.qid == qid).first().name
#         re.append([i.content, sender])
#     s.close()
#     return re  # 返回的是[内容，发送人]


def search_for_sen(word):
    """
:return
    """
    s = create_session()
    all = s.query(Questions).filter(Questions.title.like(word)).all()
    return all


def generate_talent_tabel():
    """
:return
    """
    # 日期和错题记录
    pass


def search_star_questions(page, name):  #
    """

    :param page:
    :return:
    """
    # 分页显示，每页十条
    s = create_session()
    id = s.query(Stus).filter(Stus.name == name).first().uid
    temps = s.query(Star_stu).filter(Star_stu.uid == id).limit(10).offset((page - 1) * 10).all()
    questions = []
    for i in temps:
        questions.append(s.query(Questions).filter(Questions.qid == i.qid).first())
    s.close()
    # 返回值是存有问题的数组
    return questions


def drop_and_create():
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    print('drop and create')


def do_question(qid, user_name, answer, gap):  # 题目id;是否正确
    s = create_session()
    uid = s.query(Stus).filter(Stus.name == user_name).first().uid
    # 判单正确与否
    mytype = s.query(Questions).filter(Questions.qid == qid).first().type
    myanswer = s.query(Questions).filter(Questions.qid == qid).first().answer
    mygap = s.query(Questions).filter(Questions.qid == qid).first().gap
    ques = s.query(Questions).filter(Questions.qid == qid).first()
    ques.total = ques.total + 1
    right = (mytype == 0 or mytype == 2) and answer == myanswer or mytype == 1 and gap == mygap
    print([mytype, myanswer, mygap])
    print([right, answer, gap])
    if right == 1:
        ques.right = ques.right + 1
    records = s.query(Records).filter(Records.uid == uid, Records.qid == qid).all()
    last = 0
    for i in records:
        if i.rank > last:
            last = i.rank
    new = Records(uid=uid, qid=qid, right=right, rate=0, never=0, rank=last + 1)
    s.add(new)
    s.commit()
    # 更新正确率并返回
    records = s.query(Records).filter(Records.uid == uid, Records.qid == qid).all()
    true = 0
    for j in records:
        if j.right == 1:
            true += 1
    total = last + 1
    these = s.query(Records).filter(Records.uid == uid, Records.qid == qid).all()
    for i in these:
        i.rate = true / total
    # 最近三次都做对则never为True，否则为False
    recent = s.query(Records).filter(Records.uid == uid, Records.qid == qid).order_by(desc(Records.time)).all()
    if len(recent) >= 3 and recent[0].right == 1 and recent[1].right == 1 and recent[3].right == 1:
        nevers = s.query(Records).filter(Records.uid == uid, Records.qid == qid).all()
        for i in nevers:
            i.never = 1
    else:
        nevers = s.query(Records).filter(Records.uid == uid, Records.qid == qid).all()
        for i in nevers:
            i.never = 0
    lis = [right, myanswer, mygap, int(100 * true / total), int(100 * ques.right / ques.total)]
    print([ques.right, ques.total])
    s.commit()
    s.close()
    return lis
    # 返回值为 是否正确（1为正确，0为错误） 选择题标准答案 填空题标准答案 本题本人正确率 本题整体正确率


# 形成个性化题组
def personalized_recommendation(qnum, chapters_name, choose, gap, user_name):
    # eg.(12,[2,3,4],1,0,168) means 根据168用户的错题记录，生成2、3、4、5章节的12道选择题组
    s = create_session()
    uid = s.query(Stus).filter(Stus.name == user_name).first().uid
    # 筛选出Records中 该人 该章节 该提醒 的所有错题记录
    records = s.query(Records).filter(Records.uid == uid).filter(Records.never == 0).all()
    ques = []
    for i in records:
        id = i.qid
        question = s.query(Questions).filter(Questions.qid == id).first()
        if (id not in ques) and question.chapter in chapters_name \
                and (question.type == 1 - choose or question.type == gap):
            ques.append(id)
    for i in range(len(ques)):
        for j in range(i + 1, len(ques)):
            a = s.query(Records).filter(Records.uid == uid).filter(Records.qid == ques[i]).first().rate
            b = s.query(Records).filter(Records.uid == uid).filter(Records.qid == ques[j]).first().rate
            if a > b:
                ques[i], ques[j] = ques[j], ques[i]
    # 用户做过的题里已经足够生成qnum大小的题组了
    if qnum <= len(ques):
        s.commit()
        s.close()
        return ques[0:qnum]

    # 在用户没做过Questions里选出出错率比较高的补齐
    else:
        stu = s.query(Stus).filter(Stus.name == user_name).first()
        qgroupids = [i.gid for i in stu.qgroups]
        # todo = s.query(Questions).filter(Questions.chapter in chapters_name).filter(Questions.type == 1 - choose,
        #                                                                               Questions.type == gap).all()
        todo1 = s.query(Questions).filter(Questions.chapter.in_(chapters_name)).filter(
            or_(Questions.type == 1 - choose, Questions.type == gap)).all()
        todo2 = []
        for i in todo1:
            flag = False
            for j in i.qgroups:
                if j.gid in qgroupids:
                    flag = True
                    break
            if i.uid == uid or i.public == True or flag == True:
                todo2.append(i)
        # 排序
        for i in range(len(todo2)):
            for j in range(i + 1, len(todo2)):
                if todo2[i].total == 0 and todo2[j].total != 0:
                    todo2[i], todo2[j] = todo2[j], todo2[i]
                elif todo2[i].total != 0 and todo2[j].total != 0:
                    if todo2[i].right / todo2[i].total > todo2[j].right / todo2[j].total:
                        todo2[i], todo2[j] = todo2[j], todo2[i]
        for q in todo2:
            if len(ques) == qnum:
                s.commit()
                s.close()
                return ques
            if q.qid not in ques:
                ques.append(q.qid)
    # 返回问题id


# s = create_session()
#     # 搜索范围包括questions中的public或上传者为本人的，和qgroup中的
#     uid = s.query(Stus).filter(Stus.name == user_name).first().uid
#     stu = s.query(Stus).filter(Stus.name == user_name).first()
#     qgroupids = [i.gid for i in stu.qgroups]
#     q = s.query(Questions).filter(Questions.chapter in chapters_name).filter(Questions.type == mytype).filter(
#         Questions.title.like('%' + ques_name + '%')).all()
#     ques = []
#     for i in q:
#         flag = False
#         for j in i.qgroups:
#             if j.gid in qgroupids:
#                 flag = True
#                 break
#         if i.uid == uid or i.public == True or flag == True:
#             ques.append(i)

# 返回值是Records行

# 根据id 返回 title, type, answer1, answer2, answer3, answer4
def get_question(qid):
    s = create_session()
    ques = s.query(Questions).filter(Questions.qid == qid).first()
    if ques.type == 1:
        lis = [ques.title, ques.type,
               ques.gap, ques.answerA, ques.answerB, ques.answerC, ques.answerD]
    else:
        lis = [ques.title, ques.type,
               ques.answer, ques.answerA, ques.answerB, ques.answerC, ques.answerD]
    s.commit()
    s.close()
    return lis


def getMotto(name):
    s = create_session()
    print(name)
    motto = s.query(Stus).filter(Stus.name == name).first().quote
    return motto


def get_accurate_rate(user_name):  # 查record，每章做题数，每章正确率[[],[],[]]
    s = create_session()
    uid = s.query(Stus).filter(Stus.name == user_name).first().uid
    ans = []
    for chapter in ['Chapter_1', 'Chapter_2', 'Chapter_3', 'Chapter_4', 'Chapter_5', 'Chapter_6', 'Chapter_7']:
        records = s.query(Records).filter(Records.uid == uid).all()
        total, right = 0, 0
        for i in records:
            id = i.qid
            cha = s.query(Questions).filter(Questions.qid == id).first().chapter
            if chapter == cha:
                total += 1
                if i.right == 1:
                    right += 1
        if not total == 0:
            ans.append([total, right / total])
        else:
            ans.append([0, 0.0])

    s.commit()
    s.close()
    return ans


def star_questioin(user_name, qid):
    s = create_session()
    stu = s.query(Stus).filter(Stus.name == user_name).first()
    question = s.query(Questions).filter(Questions.qid == qid).first()
    stu.starquestions.append(question)
    s.commit()
    s.close()


def get_starquestion(user_name):
    s = create_session()
    stu = s.query(Stus).filter(Stus.name == user_name).first()
    ques = []
    for i in stu.starquestions:
        ques.append(i.qid)
    return ques


def draft(ques_name):  # 关键词，章节，题型

    s = create_session()
    q = s.query(Questions).filter(Questions.title.like('%' + ques_name + '%')).first()
    lis = [q.qid, q.type, q.chapter]
    s.commit()
    s.close()
    return lis

# 用户创建了哪些问题组
def search_qgroups(user):
    s = create_session()
    uid = s.query(Stus).filter(Stus.name == user).first().uid
    groups = s.query(QGroups).filter(QGroups.uid == uid).all()
    names = []
    public = s.query(QGroups).filter(QGroups.public == True).all()
    pri = s.query(QGroups).filter(QGroups.public == False).all()
    private = []
    for i in pri:
        private.append(i)
    for i in groups:
        if not i in private:
            names.append(i.name)
    for i in public:
        if not i.name in names:
            names.append(public)


    s.close()
    s.commit()
    return names

#
def set_qgroup_public(name):
    s = create_session()
    group = s.query(QGroups).filter(QGroups.name == name).first()
    group.public = True
    s.commit()
    s.close()


def set_qgroup_private(name):
    s = create_session()
    group = s.query(QGroups).filter(QGroups.name == name).first()
    group.public = False
    s.commit()
    s.close()


def search_ques(key):
    s = create_session()
    all = s.query(Questions).all()
    name = []
    for i in all:
        if key in i.title:
            name.append(i.title)
    s.close()
    return name


def ques_in_qgroup(name):#传入的是组名
    s = create_session()
    g = s.query(QGroups).filter(QGroups.name == name).first()
    gid = g.gid
    qids_ = s.query(Ques_qgroup).filter(Ques_qgroup.gid == gid).all()
    qids = []
    for i in qids_:
        title = s.query(Questions).filter(Questions.qid == i.qid).first().title
        qids.append([i.qid, title])
    s.close()
    return qids


if __name__ == '__main__':
    # print(get_question(1))
    # print(do_question(1, "RRRR", "", "1903年11月17日；英国"))
    # print(get_question(2))
    # print(do_question(2, "RRRR", "1000", ""))
    # print(get_question(3))
    # print(do_question(3, "RRRR", "",  '2008年11月28日'))
    # print(get_question(4))
    # print(do_question(4, "RRRR", "0100", ''))
    # scope_questions("空中客车", ["Chapter_1", "Chapter_2", "
    # Chapter_3", "Chapter_4", "Chapter_5", "Chapter_6", "Chapter_7",
    #                          "Chapter_8"], 1,
    #                 "RRRR")

    Base.metadata.create_all(engine)  # 一键在数据库生成所有的类
    # change_quote("new", "fmy")
    # Base.metadata.drop_all(engine)
    # create_new_user('fmy','',True)
    # create_new_group('group', 'fmy')
    # load_files("D:\\Users\\23673\\Desktop\\summer_python\\upload_file1.xlsx", "lyj")
    # print(draft("选择题，答案是A"))
    # print(scope_questions("选择题，答案是A", ["Chapter_1"], 0, "lyj"))
    # for i in scope_questions("选择题，答案是A", ["Chapter_1"], 0, "lyj"):
    #     print(get_question(i))
    print(draft("1.3＜Ma≤5.0"))
    # change_quote("yeah", "manager")
    # s = create_session()
    # questions = s.query(Questions).filter(Questions.uid == 21371321).all()
    # for i in questions:
    #     print(i.qid)
    # s.commit()
    # s.close()
    # print(personalized_recommendation(5, ["Chapter_1", "Chapter_2"], 1, 1, "manager"))
    # Base.metadata.drop_all(engine)#一键清除S
    ###########################
    # 单题测试
    # s = create_session()
    # new = Stus(uid=21371321, name="manager")  # 此人为管理员，作为初始题目的上传者
    # s.add(new)
    # for i in range(1, 9):
    #     s.add(Chapters(name='Chapter_' + str(i), ques=[]))
    # s.commit()
    # s.close()
    # load_one_question('2008年09月28日，欧洲空中客车的A-320飞机在中国_____ 的总装公司投产。', '0001',
    #                   'Chapter_1', 1, '北京', '西安', '上海', '天津', '', True, 'manager')
    # print(get_question(1))
    # print(get_question(3))
    # print(scope_questions("莱特", ["Chapter_1", "Chapter_2"], 1, "fmy"))
    # print(do_question(1, "manager", "0001", ""))
    ################################
    # initial_data()
    # create_new_user("manager", "666666", 1)
    # print(do_question(1, "manager", "0010", ""))
    # print(do_question(2, "manager", "", "1903年12月17日；美国"))
    # print(do_question(1, "manager", "0001", ""))
    # load_one_question(title='hhh',answer=)
    # user_add_into_group(['123', 'hhhhh'], 'stu9')  # 用户主动申请加入
